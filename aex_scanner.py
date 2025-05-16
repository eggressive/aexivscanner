#!/usr/bin/env python3
"""
AEX DCF Scanner - Calculates valuation metrics and ranks AEX stocks by undervaluation
"""

import os
import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime
import time
import openpyxl
from openpyxl.styles import PatternFill, Font
import logging
import random
import requests
from requests.exceptions import HTTPError, ConnectionError, Timeout, RequestException

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('aex_scanner.log')
    ]
)
logger = logging.getLogger(__name__)

# AEX stock tickers
AEX_TICKERS = [
    'ADYEN.AS', 'ASML.AS', 'AD.AS', 'AKZA.AS', 'ABN.AS', 'DSM.AS', 'HEIA.AS', 
    'IMCD.AS', 'INGA.AS', 'KPN.AS', 'NN.AS', 'PHIA.AS', 'RAND.AS', 'REN.AS', 
    'WKL.AS', 'URW.AS', 'UNA.AS', 'MT.AS', 'RDSA.AS', 'RELX.AS', 'PRX.AS'
]

# Dictionary to store fair value estimates - add your own estimates from SimplyWall.st
# These are fair value estimates sourced from SimplyWall.st where available
FAIR_VALUE_ESTIMATES = {
    'ADYEN.AS': 1868.553,
    'ASML.AS': 768.84375,
    'AD.AS': 36.37765,
    'AKZA.AS': 69.0,
    'ABN.AS': 20.50625,
    'HEIA.AS': 91.08273,
    'IMCD.AS': 158.86667,
    'INGA.AS': 19.55111,
    'KPN.AS': 4.00857,
    'NN.AS': 53.75938,
    'PHIA.AS': 25.45814,
    'RAND.AS': 40.86875,
    'REN.AS': 51.4,
    'WKL.AS': 163.95833,
    'UNA.AS': 62.25,
    'MT.AS': 31.092327,
    'PRX.AS': 54.25765,
}

class AEXScanner:
    def __init__(self):
        """Initialize AEX Scanner"""
        # Create outputs directory if it doesn't exist
        self.output_dir = "outputs"
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Set output file path in the outputs directory
        filename = f"aex_stock_valuation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.output_file = os.path.join(self.output_dir, filename)
        
        self.data = {}
        # For rate limiting handling
        self.max_retries = 3
        self.retry_delay = 2  # seconds
        self.required_fields = ['name', 'price', 'shares_outstanding', 'fair_value']
        self.validation_results = {'success': [], 'incomplete': [], 'failed': []}
    
    def fetch_with_retry(self, ticker):
        """Fetch data for a ticker with retry logic for rate limiting"""
        retries = 0
        while retries <= self.max_retries:
            try:
                stock = yf.Ticker(ticker)
                info = stock.info
                
                # Check if response seems valid
                if not info or not isinstance(info, dict) or 'shortName' not in info:
                    if retries < self.max_retries:
                        retries += 1
                        wait_time = self.retry_delay * (1 + random.random())  # Add jitter
                        logger.warning(f"Incomplete data for {ticker}. Retrying in {wait_time:.2f}s (attempt {retries}/{self.max_retries})")
                        time.sleep(wait_time)
                        continue
                    else:
                        raise ValueError("Received incomplete data after maximum retries")
                
                # Extract relevant data
                data = {
                    'name': info.get('shortName', ticker),
                    'price': info.get('currentPrice', None),
                    'shares_outstanding': info.get('sharesOutstanding', None),
                    'fair_value': FAIR_VALUE_ESTIMATES.get(ticker, None)
                }
                
                return data
                
            except (ConnectionError, Timeout, HTTPError) as e:
                if retries < self.max_retries:
                    retries += 1
                    wait_time = self.retry_delay * (2 ** retries) * (1 + random.random())  # Exponential backoff with jitter
                    logger.warning(f"Rate limit or connection error for {ticker}: {str(e)}. Retrying in {wait_time:.2f}s (attempt {retries}/{self.max_retries})")
                    time.sleep(wait_time)
                else:
                    logger.error(f"Failed to fetch data for {ticker} after {self.max_retries} retries: {str(e)}")
                    raise
            except Exception as e:
                logger.error(f"Unexpected error fetching data for {ticker}: {str(e)}")
                raise
        
        return None
        
    def validate_stock_data(self, ticker, data):
        """Validate stock data to check for missing or incomplete fields"""
        if data is None:
            self.validation_results['failed'].append(ticker)
            return False
            
        missing_fields = [field for field in self.required_fields if field not in data or data[field] is None]
        
        if missing_fields:
            self.validation_results['incomplete'].append((ticker, missing_fields))
            logger.warning(f"Incomplete data for {ticker}: Missing {', '.join(missing_fields)}")
            return False
        else:
            self.validation_results['success'].append(ticker)
            return True
    
    def fetch_stock_data(self):
        """Fetch current stock data from Yahoo Finance with enhanced error handling"""
        logger.info("Fetching stock data from Yahoo Finance...")
        
        # Reset validation results
        self.validation_results = {'success': [], 'incomplete': [], 'failed': []}
        
        for ticker in AEX_TICKERS:
            try:
                # Fetch data with retry mechanism
                stock_data = self.fetch_with_retry(ticker)
                
                # Validate the data
                if self.validate_stock_data(ticker, stock_data):
                    self.data[ticker] = stock_data
                    logger.info(f"Successfully fetched and validated data for {ticker}: {stock_data['name']}")
                else:
                    logger.warning(f"Skipping {ticker} due to validation failure")
                
            except Exception as e:
                logger.error(f"Error processing {ticker}: {str(e)}")
                self.validation_results['failed'].append(ticker)
        
        # Log summary of fetch operation
        total_tickers = len(AEX_TICKERS)
        success_count = len(self.validation_results['success'])
        incomplete_count = len(self.validation_results['incomplete'])
        failed_count = len(self.validation_results['failed'])
        
        logger.info(f"Data fetch summary: {success_count}/{total_tickers} successful, "
                   f"{incomplete_count} incomplete, {failed_count} failed")
        
        if incomplete_count > 0:
            logger.warning("Tickers with incomplete data: " + 
                          ", ".join([t[0] for t in self.validation_results['incomplete']]))
        
        if failed_count > 0:
            logger.warning("Tickers that failed completely: " + 
                          ", ".join(self.validation_results['failed']))
    
    def calculate_metrics(self):
        """Calculate valuation metrics for each stock"""
        logger.info("Calculating valuation metrics...")
        
        for ticker, stock_data in self.data.items():
            try:
                price = stock_data['price']
                shares = stock_data['shares_outstanding']
                fair_value = stock_data['fair_value']
                
                # Calculate metrics
                market_cap = price * shares
                fair_market_cap = fair_value * shares
                discount_margin = fair_market_cap - market_cap
                discount_percent = (discount_margin / fair_market_cap) * 100 if fair_market_cap != 0 else 0
                
                # Update data dictionary with calculated metrics
                self.data[ticker].update({
                    'market_cap': market_cap,
                    'fair_market_cap': fair_market_cap,
                    'discount_margin': discount_margin,
                    'discount_percent': discount_percent
                })
                
                logger.info(f"Calculated metrics for {ticker}: Discount %: {discount_percent:.2f}%")
                
            except Exception as e:
                logger.error(f"Error calculating metrics for {ticker}: {str(e)}")
                # Mark this ticker as having calculation errors
                if ticker not in self.validation_results['failed']:
                    self.validation_results['failed'].append(ticker)
                # Remove from data dictionary to prevent downstream errors
                if ticker in self.data:
                    del self.data[ticker]
    
    def rank_stocks(self):
        """Rank stocks by discount percentage (undervaluation)"""
        logger.info("Ranking stocks by discount percentage...")
        
        # Convert to DataFrame for easier manipulation
        df = pd.DataFrame.from_dict(self.data, orient='index')
        
        # Sort by discount percentage (descending)
        df = df.sort_values(by='discount_percent', ascending=False)
        
        return df
    
    def save_to_excel(self, df):
        """Save results to Excel file with formatting"""
        logger.info(f"Saving results to {self.output_file}...")
        
        # Create Excel writer
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # Convert to millions for market cap values
            df['market_cap_M'] = df['market_cap'] / 1_000_000
            df['fair_market_cap_M'] = df['fair_market_cap'] / 1_000_000
            df['discount_margin_M'] = df['discount_margin'] / 1_000_000
            
            # Select and rename columns for output
            output_df = df[[
                'name', 'price', 'fair_value', 
                'market_cap_M', 'fair_market_cap_M', 
                'discount_margin_M', 'discount_percent'
            ]].rename(columns={
                'name': 'Company',
                'price': 'Current Price',
                'fair_value': 'Fair Value',
                'market_cap_M': 'Market Cap (M)',
                'fair_market_cap_M': 'Fair Market Cap (M)',
                'discount_margin_M': 'Discount Margin (M)',
                'discount_percent': 'Discount %'
            })
            
            # Format numbers
            for col in ['Market Cap (M)', 'Fair Market Cap (M)', 'Discount Margin (M)']:
                output_df[col] = output_df[col].map('{:,.2f}'.format)
                
            output_df['Current Price'] = output_df['Current Price'].map('{:.2f}'.format)
            output_df['Fair Value'] = output_df['Fair Value'].map('{:.2f}'.format)
            output_df['Discount %'] = output_df['Discount %'].map('{:.2f}%'.format)
            
            # Write to Excel with startrow=2 to leave room for title and descriptions
            output_df.to_excel(writer, sheet_name='AEX Valuation', index=True, index_label='Ticker', startrow=2)
            
            # Apply formatting
            workbook = writer.book
            worksheet = writer.sheets['AEX Valuation']
            
            # Add a header with timestamp in row 1
            worksheet['A1'] = f"AEX Stock Valuation Scanner - Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            worksheet.merge_cells('A1:H1')
            header_font = Font(bold=True, size=14)
            worksheet['A1'].font = header_font
            
            # Add column descriptions in row 2
            descriptions = {
                'A': 'Ticker: Stock ticker symbol (Amsterdam Exchange)',
                'B': 'Company: Company name',
                'C': 'Current Price: Current stock price in Euros',
                'D': 'Fair Value: Estimated fair value per share based on DCF analysis',
                'E': 'Market Cap (M): Current market capitalization in millions',
                'F': 'Fair Market Cap (M): Fair market capitalization based on fair value in millions',
                'G': 'Discount Margin (M): Difference between fair and current market cap in millions',
                'H': 'Discount %: How undervalued/overvalued the stock is'
            }
            
            # Insert descriptions in row 2
            for col, description in descriptions.items():
                worksheet[f'{col}2'] = description
                worksheet[f'{col}2'].font = Font(italic=True, size=8)
                
            # Set column widths to fit descriptions
            for col in 'ABCDEFGH':
                worksheet.column_dimensions[col].width = 20
                    
            # Add conditional formatting to highlight undervalued/overvalued stocks
            # Create fill colors
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')    # Light red
            
            # Apply color formatting to the Discount % column
            discount_col = 'H'
            # Data starts at row 3 (after title and descriptions)
            for row_num in range(3 + 1, worksheet.max_row + 1):  # +1 to skip header row
                cell = worksheet[f"{discount_col}{row_num}"]
                if cell.value:
                    # Extract numeric value from formatted string (remove % sign and convert to float)
                    try:
                        value_str = cell.value.replace('%', '').strip()
                        value = float(value_str)
                        if value >= 10:  # Undervalued by 10% or more
                            cell.fill = green_fill
                        elif value <= -10:  # Overvalued by 10% or more
                            cell.fill = red_fill
                    except (ValueError, AttributeError):
                        pass  # Skip if can't convert to float
            
        logger.info(f"Results saved to {self.output_file}")
        
    def run(self):
        """Run the full scanning process"""
        logger.info("Starting AEX DCF Scanner...")
        
        self.fetch_stock_data()
        
        if not self.data:
            logger.error("No valid stock data was fetched. Aborting scan.")
            return None
            
        self.calculate_metrics()
        ranked_df = self.rank_stocks()
        
        if ranked_df.empty:
            logger.error("No stocks with valid metrics to rank. Aborting scan.")
            return None
            
        self.save_to_excel(ranked_df)
        
        # Generate a comprehensive summary
        total_tickers = len(AEX_TICKERS)
        processed_tickers = len(self.data)
        skipped_tickers = total_tickers - processed_tickers
        
        summary = (
            f"Scan completed successfully!\n"
            f"Processed {processed_tickers}/{total_tickers} tickers successfully.\n"
            f"Skipped {skipped_tickers} tickers due to data issues."
        )
        
        logger.info(summary)
        return self.output_file

if __name__ == "__main__":
    scanner = AEXScanner()
    output_file = scanner.run()
    print(f"Scan complete! Results saved to: {output_file}")
